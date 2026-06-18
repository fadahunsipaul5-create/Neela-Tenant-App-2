"""
Import income statement data from the Neela Capital Investments Excel workbook.

Run: python manage.py import_income_statement
      python manage.py import_income_statement --clear
      python manage.py import_income_statement --dry-run
"""
import re
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction

from api.models import OperatingExpense, Payment, Property, PropertyFinancials, Tenant

IMPORT_TAG = 'excel-import-2026'
EXCEL_NAME = '2026 Neela Capital Invesments LLC..xlsx'

SHEET_TO_KEY = {
    'Bella_Jess': 'Bella Jess',
    'Tomball': 'Tomabll',
    'Conroe': 'Conroe',
    'Ave_Q': 'Ave Q',
    'Sherman': 'Sherman',
    '70th': '70th',
    'Avenue H': 'Ave H',
    'Wooding': 'Wooden',
    'Avenue F': 'Ave F',
}

MONTH_LABELS = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
}

FINANCIAL_LABELS = {
    'purchase price': 'purchase_price',
    'purchase price (2024)': 'purchase_price',
    'purchase price (2025)': 'purchase_price',
    'down payment': 'down_payment',
    'closing cost': 'closing_cost',
    'loan amount': 'loan_amount',
    'interest rate': 'interest_rate',
    'loan term (years)': 'loan_term_years',
    'monthly mortgage payment (p&i)': 'monthly_mortgage_payment',
    'annual depreciation': 'annual_depreciation_years',
    'annual depreciation (27.5 yrs)': 'annual_depreciation_years',
}


def project_root():
    return Path(__file__).resolve().parent.parent.parent.parent.parent


def excel_path():
    return project_root() / EXCEL_NAME


def normalize(text):
    return re.sub(r'[^a-z0-9]+', '', (text or '').lower())


def parse_decimal(val):
    if val is None:
        return None
    if isinstance(val, (int, float, Decimal)):
        try:
            return Decimal(str(val))
        except InvalidOperation:
            return None
    s = str(val).strip()
    if not s or s in ('-', '#REF!', '#N/A', 'None'):
        return None
    s = s.replace(',', '').replace('$', '')
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def parse_loan_term(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip().lower()
    m = re.search(r'(\d+)', s)
    return int(m.group(1)) if m else None


def map_category(label):
    l = label.lower()
    if 'management' in l:
        return 'management'
    if 'repair' in l or 'maintenance' in l or 'inspection' in l or 'appraisal' in l:
        return 'maintenance'
    if 'tax' in l:
        return 'taxes'
    if 'insurance' in l:
        return 'insurance'
    if 'hoa' in l:
        return 'hoa'
    if 'advertis' in l or 'leasing' in l:
        return 'advertising'
    if 'legal' in l or 'professional' in l or 'survey' in l:
        return 'legal'
    if 'supplies' in l or 'materials' in l:
        return 'supplies'
    if 'transport' in l or 'mile' in l or 'milleage' in l or 'trip' in l:
        return 'transportation'
    if 'bank' in l:
        return 'bank_charges'
    if 'mortgage' in l or 'p&i' in l:
        return 'mortgage_principal'
    if 'depreciation' in l:
        return 'depreciation'
    if any(x in l for x in ('electric', 'water', 'internet', 'utilities', 'gas')):
        return 'utilities'
    if 'clean' in l:
        return 'cleaning'
    return 'other'


def is_admin_expense(label, category):
    l = label.lower()
    if category in ('mortgage_principal', 'mortgage_interest', 'depreciation'):
        return True
    if 'mortgage' in l or 'depreciation' in l:
        return True
    return False


def parse_overview(wb):
    ws = wb['Neela Real Estate ']
    rows = list(ws.iter_rows(values_only=True))
    names = rows[2]
    addrs = rows[3]
    props = {}
    for i, name in enumerate(names):
        if not name or str(name).strip() in ('CATEGORY', ''):
            continue
        key = str(name).strip()
        if not key or key == 'CATEGORY':
            continue
        props[key] = {
            'name': key,
            'address': str(addrs[i]).strip() if i < len(addrs) and addrs[i] else '',
            'units': 1,
            'purchase_price': None,
            'down_payment': None,
            'closing_cost': None,
            'loan_amount': None,
            'interest_rate': None,
            'loan_term_years': None,
            'monthly_mortgage_payment': None,
            'annual_depreciation_years': Decimal('27.5'),
            'escrow_notes': '',
        }
    field_map = {
        '# of units': 'units',
        'purchase price': 'purchase_price',
        'down payment': 'down_payment',
        'closing cost': 'closing_cost',
        'loan amount': 'loan_amount',
        'interest rate': 'interest_rate',
        'loan term (years)': 'loan_term_years',
        'monthly mortgage payment (p&i)': 'monthly_mortgage_payment',
        'annual depreciation': 'annual_depreciation_years',
    }
    for row in rows[4:30]:
        if len(row) < 5 or not row[4]:
            continue
        label = str(row[4]).strip().lower()
        if label not in field_map:
            continue
        field = field_map[label]
        for key, meta in props.items():
            col = None
            for i, n in enumerate(names):
                if n and str(n).strip() == key:
                    col = i
                    break
            if col is None or col >= len(row):
                continue
            val = row[col]
            if field == 'units':
                meta['units'] = int(parse_decimal(val) or 1)
            elif field == 'loan_term_years':
                meta['loan_term_years'] = parse_loan_term(val)
            elif field == 'annual_depreciation_years':
                meta['annual_depreciation_years'] = parse_decimal(val) or Decimal('27.5')
            elif field == 'interest_rate':
                d = parse_decimal(val)
                if d is not None:
                    if d > 1:
                        d = d / Decimal('100')
                    if d > Decimal('1'):
                        d = None
                meta['interest_rate'] = d
            else:
                meta[field] = parse_decimal(val)
    return props


def find_property(key, meta):
    keywords = [normalize(meta['name']), normalize(meta['address'])]
    keywords = [k for k in keywords if k]
    best = None
    best_score = 0
    for prop in Property.objects.all():
        blob = normalize(f"{prop.name} {prop.address} {prop.area or ''}")
        score = sum(1 for k in keywords if k in blob)
        if 'unit' in prop.name.lower():
            score -= 0.5
        if score > best_score:
            best_score = score
            best = prop
    if best and best_score >= 1:
        return best
    street = meta['address'].split(' Houston')[0].strip() if meta['address'] else meta['name']
    return Property.objects.create(
        name=meta['name'],
        address=street or meta['name'],
        city='Houston',
        state='Texas',
        units=meta.get('units') or 1,
        area=meta['name'],
    )


def get_import_tenant(prop):
    email = f"excel-import-{prop.id}@neela.local"
    tenant, _ = Tenant.objects.get_or_create(
        email=email,
        defaults={
            'name': f'Rent Roll — {prop.name}',
            'phone': '0000000000',
            'status': 'Active',
            'property_unit': prop.address,
            'rent_amount': Decimal('0'),
            'deposit': Decimal('0'),
        },
    )
    if tenant.property_unit != prop.address:
        tenant.property_unit = prop.address
        tenant.save(update_fields=['property_unit'])
    return tenant


def parse_property_sheet(ws, year):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return {}, {}, {}

    month_cols = []
    header = rows[1] if len(rows) > 1 else []
    for i, val in enumerate(header):
        if hasattr(val, 'month') and hasattr(val, 'year'):
            month_cols.append((val.month, i))

    summaries = {}
    for row in rows[2:20]:
        if not row or len(row) < 5 or not row[1]:
            continue
        label = str(row[1]).strip().lower().rstrip('.')
        month = MONTH_LABELS.get(label)
        if not month:
            continue
        summaries[month] = {
            'income': parse_decimal(row[2]),
            'expenses': parse_decimal(row[3]),
            'noi': parse_decimal(row[4]),
        }

    monthly_expenses = {m: [] for m in range(1, 13)}
    sheet_financials = {}

    for month, col in month_cols:
        in_operating = False
        for row in rows[4:]:
            if col >= len(row):
                continue
            cat = row[col]
            amt = row[col + 1] if col + 1 < len(row) else None
            if cat is None:
                continue
            cat_s = str(cat).strip()
            if not cat_s or cat_s == 'Category':
                continue
            if cat_s.upper() == 'OPERATING EXPENSES':
                in_operating = True
                continue
            if cat_s == 'PROPERTY OVERVIEW':
                in_operating = False
                if month == 1:
                    fin_key = cat_s.lower()
                    _ = fin_key
                continue
            if 'total operating expenses' in cat_s.lower():
                break
            if not in_operating:
                fin_field = FINANCIAL_LABELS.get(cat_s.lower())
                if fin_field and month == 1:
                    if fin_field == 'loan_term_years':
                        sheet_financials[fin_field] = parse_loan_term(amt)
                    elif fin_field == 'interest_rate':
                        d = parse_decimal(amt)
                        if d and d > 1:
                            d = d / Decimal('100')
                        sheet_financials[fin_field] = d
                    elif fin_field == 'annual_depreciation_years':
                        sheet_financials[fin_field] = parse_decimal(amt) or Decimal('27.5')
                    else:
                        sheet_financials[fin_field] = parse_decimal(amt)
                continue
            amount = parse_decimal(amt)
            if amount is None or amount <= 0:
                continue
            if cat_s.lower().startswith('$ per mile') or 'number of trips' in cat_s.lower():
                continue
            monthly_expenses[month].append((cat_s, amount))

    return summaries, monthly_expenses, sheet_financials


def sanitize_financials(data):
    ir = data.get('interest_rate')
    if ir is not None:
        if ir > 1:
            ir = ir / Decimal('100')
        if ir > Decimal('1') or ir < 0:
            data['interest_rate'] = None
        else:
            data['interest_rate'] = ir
    loan = data.get('loan_amount')
    if loan is not None and loan > Decimal('10000000'):
        data['loan_amount'] = None
    term = data.get('loan_term_years')
    if term is not None and term > 50:
        data['loan_term_years'] = None
    return data


def upsert_financials(prop, overview_meta, sheet_financials):
    data = {
        'purchase_price': sheet_financials.get('purchase_price') or overview_meta.get('purchase_price'),
        'down_payment': sheet_financials.get('down_payment') or overview_meta.get('down_payment'),
        'closing_cost': sheet_financials.get('closing_cost') or overview_meta.get('closing_cost'),
        'loan_amount': sheet_financials.get('loan_amount') or overview_meta.get('loan_amount'),
        'interest_rate': sheet_financials.get('interest_rate') or overview_meta.get('interest_rate'),
        'loan_term_years': sheet_financials.get('loan_term_years') or overview_meta.get('loan_term_years'),
        'monthly_mortgage_payment': sheet_financials.get('monthly_mortgage_payment') or overview_meta.get('monthly_mortgage_payment'),
        'annual_depreciation_years': sheet_financials.get('annual_depreciation_years') or overview_meta.get('annual_depreciation_years') or Decimal('27.5'),
        'escrow_notes': overview_meta.get('escrow_notes') or '',
    }
    data = sanitize_financials(data)
    PropertyFinancials.objects.update_or_create(property=prop, defaults=data)


class Command(BaseCommand):
    help = 'Import income statement data from the Neela Capital Investments Excel workbook'

    def add_arguments(self, parser):
        parser.add_argument('--year', type=int, default=2026)
        parser.add_argument('--clear', action='store_true', help='Remove prior excel-import-2026 rows first')
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        year = options['year']
        dry_run = options['dry_run']
        path = excel_path()
        if not path.is_file():
            self.stderr.write(self.style.ERROR(f'Excel not found: {path}'))
            return

        try:
            import openpyxl
        except ImportError:
            self.stderr.write(self.style.ERROR('Install openpyxl'))
            return

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        overview = parse_overview(wb)

        if options['clear'] and not dry_run:
            Payment.objects.filter(reference__startswith=IMPORT_TAG).delete()
            OperatingExpense.objects.filter(notes__startswith=IMPORT_TAG).delete()
            self.stdout.write('Cleared previous excel import rows.')

        stats = {'properties': 0, 'payments': 0, 'expenses': 0}

        for sheet_name, key in SHEET_TO_KEY.items():
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f'Sheet missing: {sheet_name}'))
                continue
            meta = overview.get(key)
            if not meta:
                self.stdout.write(self.style.WARNING(f'No overview for {key}'))
                continue

            summaries, monthly_expenses, sheet_financials = parse_property_sheet(wb[sheet_name], year)

            for month in range(1, 13):
                summary = summaries.get(month, {})
                income = summary.get('income')
                if income and income > 0:
                    stats['payments'] += 1
                stats['expenses'] += len(monthly_expenses.get(month, []))

            if dry_run:
                stats['properties'] += 1
                self.stdout.write(f"[dry-run] {key} ({sheet_name})")
                continue

            with transaction.atomic():
                prop = find_property(key, meta)
                tenant = get_import_tenant(prop)
                upsert_financials(prop, meta, sheet_financials)
                stats['properties'] += 1
                self.stdout.write(f"{prop.name} ({sheet_name})")

                payment_objs = []
                expense_objs = []

                for month in range(1, 13):
                    summary = summaries.get(month, {})
                    income = summary.get('income')
                    if income and income > 0:
                        payment_objs.append(Payment(
                            reference=f"{IMPORT_TAG}-{prop.id}-{year}-{month:02d}-rent",
                            tenant=tenant,
                            amount=income,
                            date=date(year, month, 1),
                            status='Paid',
                            type='Rent',
                            method='Excel Import',
                        ))

                    for label, amount in monthly_expenses.get(month, []):
                        category = map_category(label)
                        visibility = 'admin_only' if is_admin_expense(label, category) else 'operating'
                        note = f"{IMPORT_TAG}|{sheet_name}|{month:02d}|{label[:80]}"
                        expense_objs.append(OperatingExpense(
                            property=prop,
                            amount=amount,
                            category=category,
                            visibility=visibility,
                            date=date(year, month, 1),
                            notes=note,
                        ))

                for p in payment_objs:
                    Payment.objects.update_or_create(
                        reference=p.reference,
                        defaults={
                            'tenant': p.tenant,
                            'amount': p.amount,
                            'date': p.date,
                            'status': p.status,
                            'type': p.type,
                            'method': p.method,
                        },
                    )
                if expense_objs:
                    OperatingExpense.objects.bulk_create(expense_objs, batch_size=200)

        wb.close()
        self.stdout.write(self.style.SUCCESS(
            f"Done — properties: {stats['properties']}, payments: {stats['payments']}, expenses: {stats['expenses']}"
            + (' (dry run)' if dry_run else '')
        ))
